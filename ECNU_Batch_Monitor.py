import time
import os
import requests
import urllib3
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font

# 复用你原脚本中的核心逻辑
from ECNU_Courses_Crawler import init_driver, login_and_get_session, BASE_URL_TEMPLATE

# 禁用安全警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ================= 监测配置区 =================
MONITOR_CODES = [
    "MARX1001.13",
    "MARX1001.14",
    "CS1304.01",
    "CS2101.02",
    "CS2101.03",
    "CS2101.04",
    "CS2203.01",
    "CS2203.02",
    "CS2204.01",
    "CS2204.02",
    "CS2304.01",
    "CS2305.01",
    "CS2306.01",
    "CS2307.02",
    "CS2308.02",
    "CS2309.02",
    "CS2310.01",
    "CS2311.01",
    "CS3301.01",
]
INTERVAL = 60 
PAGES_TO_CHECK = 4
OUTPUT_FILE = "Course_Monitor_Log.xlsx" # 必须是 xlsx 才能保存样式
# ============================================

def parse_teachers(teacher_list):
    """根据提供的示例结构解析教师姓名"""
    if not teacher_list or not isinstance(teacher_list, list):
        return "未知"
    # 修复点：访问 person 节点下的 nameZh
    names = [t.get('person', {}).get('nameZh', '未知') for t in teacher_list]
    return ", ".join(names)

def update_excel_with_format(new_row_data):
    """将数据追加到 Excel 并自动调整样式"""
    cols = ["记录时间", "课程名称", "授课教师", "课号", "实际人数", "上限人数", "剩余名额", "选课比例"]
    df_new = pd.DataFrame([new_row_data], columns=cols)

    # 1. 追加数据
    if not os.path.exists(OUTPUT_FILE):
        df_new.to_excel(OUTPUT_FILE, index=False)
    else:
        with pd.ExcelWriter(OUTPUT_FILE, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            # 读取原有数据确定起始行
            try:
                start_row = writer.book['Sheet1'].max_row
                df_new.to_excel(writer, index=False, header=False, startrow=start_row)
            except Exception:
                df_new.to_excel(writer, index=False)

    # 2. 调整样式 (行高列宽/自动换行)
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    
    alignment_style = Alignment(wrap_text=True, vertical='center', horizontal='left')
    header_font = Font(bold=True)

    for column_cells in ws.columns:
        col_letter = column_cells[0].column_letter
        max_length = 0
        for cell in column_cells:
            cell.alignment = alignment_style # 设置自动换行
            if cell.row == 1:
                cell.font = header_font
            
            # 计算最大列宽
            try:
                val_str = str(cell.value) if cell.value else ""
                # 中文字符长度补偿计算
                content_len = len(val_str.encode('gbk')) 
                if content_len > max_length:
                    max_length = content_len
            except: pass
        
        # 自动调整列宽：最小12，最大60
        adjusted_width = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[col_letter].width = adjusted_width

    # 设置所有行高为自动（或固定高度）
    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20 # 设置一个舒适的基础行高

    try:
        wb.save(OUTPUT_FILE)
    except PermissionError:
        print(f"[!] 无法保存文件！请关闭已打开的 {OUTPUT_FILE}")

def check_and_log(session):
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[*] {now_str} 正在轮询数据...")
    
    for page in range(1, PAGES_TO_CHECK + 1):
        url = BASE_URL_TEMPLATE.format(page=page)
        try:
            resp = session.get(url, verify=False, timeout=10)
            data = resp.json()
            rows = data.get('data', data) if isinstance(data, dict) else data
            if not rows: continue

            for course in rows:
                c_lesson_code = course.get("code", "") 
                
                if c_lesson_code in MONITOR_CODES:
                    c_name = course.get("course", {}).get("nameZh", "")
                    teachers = parse_teachers(course.get("teacherAssignmentList", []))
                    std_count = course.get("stdCount", 0)
                    limit_count = course.get("limitCount", 0)
                    
                    remaining = limit_count - std_count
                    fill_ratio = f"{(std_count / limit_count * 100):.2f}%" if limit_count > 0 else "0.00%"
                    
                    # 准备数据行
                    row_data = [now_str, c_name, teachers, c_lesson_code, std_count, limit_count, remaining, fill_ratio]
                    
                    # 写入 Excel 并格式化
                    update_excel_with_format(row_data)
                    print(f"    - [写入成功] {c_lesson_code} | 教师: {teachers}")
            
        except Exception as e:
            print(f"    [!] 轮询异常: {e}")
            return "ERROR"
    return "SUCCESS"

def main():
    print(f"[*] 监测启动，结果保存至: {OUTPUT_FILE}")
    driver = init_driver()
    session = None
    try:
        while True:
            if session is None:
                session = login_and_get_session(driver)
                if not session:
                    time.sleep(60); continue

            status = check_and_log(session)
            if status == "SESSION_EXPIRED":
                session = None; continue
            
            time.sleep(INTERVAL)
    except KeyboardInterrupt:
        print("\n[*] 监测停止。")
    finally:
        driver.quit()

if __name__ == "__main__":
    main()